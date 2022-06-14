import openpyxl as xl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook(
    "D:\\Collection\\Programming Languages\\Phython\\ExcelSheet\\transactions.xlsx")
sheet = wb["Sheet1"]
sheet["d1"] = "Updated Price"
corrected_cell_name = sheet["d1"].value
# print(corrected_cell_name)
for row in range(2, sheet.max_row+1):
    cell = sheet.cell(row, 3)
    corrected_prices = cell.value * 0.9  # 10% discount
    corrected_prices_cell = sheet.cell(row, 4)
    corrected_prices_cell.value = corrected_prices
    # print("$" + str(corrected_prices))
    # print(cell.value)

values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, "b6")


wb.save("D:\\Collection\\Programming Languages\\Phython\\ExcelSheet(Transaction_one_file)\\transactions_update.xlsx")
